"""
Turning a government spreadsheet into observations.

The output is one flat list of `(factor, area_code, period, value)`, which is
all the rest of the system wants. Everything difficult is between here and
there, and it is all the same difficulty: these files are laid out for a human
reading them, not for a program.

What that means in practice, and what each rule is defending against:

* **Find the header row; never assume it.** Release notes, a title, a blank
  spacer and a logo all live above it, and the count changes between editions.
* **Find the area column by its values.** The heading is "Area code",
  "LAD24CD", "Local authority code" or "ladcode23" depending on which file you
  opened. Every one of them holds strings like `E07000209`.
* **Match value columns on a fragment, then sanity-check the numbers.** A
  fragment can match the wrong column — "median" appears in both "median rent"
  and "median week" — so every spec carries bounds, and a column whose values
  fall outside them is rejected rather than stored.
* **Refuse an empty parse.** A file that resolved, opened and yielded nothing
  is the most dangerous outcome: it looks like success and it silently empties
  a factor. It raises.
"""

import re
from typing import Dict, Iterator, List, Optional, Tuple

from ons.datasets import LAD_PREFIXES, Column, Dataset

#: E06000001, W06000015, and so on.
AREA_CODE = re.compile(r"^[EWSN]\d{8}$")
YEAR = re.compile(r"(?:^|\D)(19|20)\d{2}(?:\D|$)")
#: '2024-07', '2024 JUL', 'Jul-24'.
MONTHS = ("jan", "feb", "mar", "apr", "may", "jun",
          "jul", "aug", "sep", "oct", "nov", "dec")


class ParseError(RuntimeError):
    """The file did not have the shape the spec describes."""


Observation = Tuple[str, str, str, float]      # factor, area, period, value


def _text(cell) -> str:
    return "" if cell is None else str(cell).strip()


def _number(cell) -> Optional[float]:
    """A cell as a number, or None.

    Suppressed and unavailable values are published as `:`, `x`, `..`, `#` or
    `N/A` depending on the file and the decade. All of them mean "no
    observation", which is a gap and never a zero.
    """
    if cell is None:
        return None
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        value = float(cell)
        return None if value != value else value          # drop NaN
    text = str(cell).strip().replace(",", "").replace("£", "").replace("%", "")
    if not text or text in {":", "x", "..", ".", "-", "#", "N/A", "n/a", "c"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_header(rows: List[list], contains: str, limit: int = 30) -> int:
    """Index of the header row: the first one holding `contains`, above data.

    Searching only the first `limit` rows on purpose. A file with 40 rows of
    preamble is not a file this spec understands, and guessing further down
    risks locking onto a section heading in the middle of the data.
    """
    needle = contains.lower()
    for i, row in enumerate(rows[:limit]):
        for cell in row:
            if needle in _text(cell).lower():
                return i
    raise ParseError(
        f"no header row containing {contains!r} in the first {limit} rows; "
        f"the file's layout changed or the spec points at the wrong sheet")


def find_area_column(rows: List[list], header_row: int) -> int:
    """Index of the column holding ONS area codes, found by looking at them.

    Column headings for this differ across every file we read; the values do
    not. The column with the most cells matching the code pattern wins.
    """
    body = rows[header_row + 1:header_row + 200]
    if not body:
        raise ParseError("no data rows below the header")

    best, best_hits = -1, 0
    width = max(len(r) for r in body)
    for col in range(width):
        hits = sum(1 for row in body
                   if col < len(row) and AREA_CODE.match(_text(row[col])))
        if hits > best_hits:
            best, best_hits = col, hits

    # Half the rows below the header, not a fixed count: a real per-authority
    # table is ~300 rows of which ~300 match, and a sheet where most rows are
    # not area codes is a summary or a notes tab whichever way it is counted.
    if best_hits < max(1, len(body) // 2):
        raise ParseError(
            "no column of ONS area codes (E06000001 and similar). Either the "
            "release changed geography, or this sheet is a summary rather than "
            "the per-authority table")
    return best


def find_value_column(header: list, fragment: str) -> Optional[int]:
    """Index of the first column whose heading contains `fragment`."""
    needle = fragment.lower()
    for i, cell in enumerate(header):
        if needle in _text(cell).lower():
            return i
    return None


def normalise_period(raw: str, cadence: str) -> Optional[str]:
    """A cell's period as 'YYYY-MM' or 'YYYY'.

    ONS writes dates as '2024 JUL', '2024-07', 'Jul-24' and as a real datetime
    depending on the file, and an annual file's columns are bare years.
    """
    text = raw.strip()
    if not text:
        return None

    iso = re.match(r"^(\d{4})-(\d{2})", text)
    if iso:
        return text[:7] if cadence == "monthly" else iso.group(1)

    spaced = re.match(r"^(\d{4})\s+([A-Za-z]{3})", text)
    if spaced:
        month = MONTHS.index(spaced.group(2).lower()[:3]) + 1
        return f"{spaced.group(1)}-{month:02d}"

    short = re.match(r"^([A-Za-z]{3})[-\s](\d{2,4})$", text)
    if short and short.group(1).lower()[:3] in MONTHS:
        month = MONTHS.index(short.group(1).lower()[:3]) + 1
        year = int(short.group(2))
        year += 2000 if year < 100 else 0
        return f"{year}-{month:02d}"

    year = re.search(r"(19|20)\d{2}", text)
    if year:
        return year.group(0)
    return None


def _accept(column: Column, value: float) -> bool:
    """Bounds are how a fragment match proves it found the right column."""
    scaled = value * column.scale
    if column.lo is not None and scaled < column.lo:
        return False
    if column.hi is not None and scaled > column.hi:
        return False
    return True


def parse_sheet(rows: List[list], dataset: Dataset) -> List[Observation]:
    """Every observation in one sheet, as (factor, area, period, value)."""
    rows = [r for r in rows if any(_text(c) for c in r)]
    if not rows:
        raise ParseError("the sheet is empty")

    header_row = find_header(rows, dataset.header_contains)
    header = rows[header_row]
    area_col = find_area_column(rows, header_row)
    body = rows[header_row + 1:]

    wide_years = dataset.extra.get("wide_years") == "true"
    period_col = (find_value_column(header, dataset.period_column)
                  if dataset.period_column else None)
    if dataset.period_column and period_col is None:
        raise ParseError(
            f"no period column containing {dataset.period_column!r}; "
            f"headers were {[_text(c) for c in header][:12]}")

    out: List[Observation] = []
    rejected: Dict[str, int] = {}

    for column in dataset.columns:
        if wide_years:
            # One column per year: every heading that looks like a year, with
            # the fragment still narrowing which block of the sheet to read.
            targets = [(i, normalise_period(_text(c), "annual"))
                       for i, c in enumerate(header)
                       if YEAR.search(_text(c))]
        else:
            found = find_value_column(header, column.header_contains)
            if found is None:
                raise ParseError(
                    f"no column containing {column.header_contains!r} for "
                    f"{column.factor}; headers were "
                    f"{[_text(c) for c in header][:12]}")
            targets = [(found, None)]

        for col, fixed_period in targets:
            for row in body:
                if col >= len(row) or area_col >= len(row):
                    continue
                area = _text(row[area_col])
                if not AREA_CODE.match(area) or not area.startswith(LAD_PREFIXES):
                    continue

                value = _number(row[col])
                if value is None:
                    continue                       # a gap, never a zero
                if not _accept(column, value):
                    rejected[column.factor] = rejected.get(column.factor, 0) + 1
                    continue

                if fixed_period:
                    period = fixed_period
                elif period_col is not None:
                    period = normalise_period(_text(row[period_col]),
                                              dataset.cadence)
                else:
                    period = dataset.period
                if not period:
                    continue

                out.append((column.factor, area, period,
                            round(value * column.scale, 3)))

    # Out-of-range values are diagnosed before emptiness, because "every value
    # was out of range" and "there were no values" look identical from the
    # outside and have completely different fixes.
    for factor, n in rejected.items():
        # A handful of suppressed or odd authorities is normal. Wholesale
        # rejection means the fragment matched the wrong column.
        kept = sum(1 for f, *_ in out if f == factor)
        if n > kept:
            raise ParseError(
                f"{factor}: {n} values were outside the expected range and "
                f"only {kept} inside it. The header fragment "
                f"almost certainly matched the wrong column")

    if not out:
        raise ParseError(
            "the file parsed but produced no observations. That is the one "
            "outcome that must never pass silently — it would empty a factor "
            "while looking like a successful refresh")
    return out


def parse_workbook(path, dataset: Dataset) -> List[Observation]:
    """Open a workbook and parse the sheet the spec names."""
    import openpyxl

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if dataset.sheet and dataset.sheet in book.sheetnames:
            sheet = book[dataset.sheet]
        elif dataset.sheet:
            # Named sheets get renamed between releases — "Table 3" becomes
            # "3a". Fall back to a fuzzy match before giving up, because the
            # alternative is a failed refresh over a renamed tab.
            wanted = dataset.sheet.lower().replace(" ", "")
            match = next((n for n in book.sheetnames
                          if wanted in n.lower().replace(" ", "")), None)
            if match is None:
                raise ParseError(
                    f"no sheet named {dataset.sheet!r}; the workbook has "
                    f"{book.sheetnames}")
            sheet = book[match]
        else:
            sheet = book[book.sheetnames[dataset.sheet_index]]

        return parse_sheet([list(r) for r in sheet.iter_rows(values_only=True)],
                           dataset)
    finally:
        book.close()


def to_store(observations: List[Observation]) -> Dict[str, Dict[str, Dict[str, float]]]:
    """Observations as `{factor: {area_code: {period: value}}}`.

    The shape the runtime reads: a factor and an area are both known before a
    period is, so this nests in the order the lookups happen.
    """
    out: Dict[str, Dict[str, Dict[str, float]]] = {}
    for factor, area, period, value in observations:
        out.setdefault(factor, {}).setdefault(area, {})[period] = value
    return out


def summarise(observations: List[Observation]) -> Iterator[str]:
    by_factor: Dict[str, List[Observation]] = {}
    for obs in observations:
        by_factor.setdefault(obs[0], []).append(obs)
    for factor, rows in sorted(by_factor.items()):
        periods = sorted({r[2] for r in rows})
        areas = {r[1] for r in rows}
        yield (f"{factor}: {len(rows):,} values, {len(areas)} authorities, "
               f"{len(periods)} period(s) {periods[0]}..{periods[-1]}")
